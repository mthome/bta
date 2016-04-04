#!/usr/bin/env python

from __future__ import print_function
import myfitnesspal
import sys
from datetime import timedelta, datetime, date
from openpyxl import Workbook
from openpyxl.comments import Comment
from argparse import ArgumentParser

oneday = timedelta(days=1)

class Record(object):
    def __init__(self, date, **kwargs):
        self.date = date
        for arg,v in kwargs.items():
            self.__setattr__(arg,v)
    def grok_notes(self, notes):
        self.notes = notes
        '''
fasting 7:30am 101
bkfst 7:50am NR
coffee 8:30am  11:30am 97
amsnack 12:20pm NR
lunch 1:15pm NR 2:30pm 112 3:30pm 110
afsnack 3:45pm  NR 5:50pm 144
dinner 8:15pm 129
AB 11;45pm 123
        '''
        lines = notes.split('\n')
        m = 'unknown'
        #for line in lines:
        #print('\n{}'.format(notes))

class BTA(object):
    def __init__(self,username):
        self.username = username
        self.records = list()
        self.verbose = 0
        
    def get_data(self, d1, d2):
        client = myfitnesspal.Client(self.username)
        if self.verbose:
            print('getting data for {} - {}'.format(d1,d2))
        if self.verbose:
            print('getting weights')
        self.weights = client.get_measurements('Weight', d1, d2)
        if self.verbose:
            print('getting steps')
        self.steps = client.get_measurements('Fitbit steps', d1, d2)
        weight = 0
        
        d = d1
        while d <= d2:
            if self.verbose:
                print('getting details for {}'.format(d))
            
            carbs = 0
            cals = 0
            meals = {}
            notes = ''
            try:
                weight = self.weights[d]
            except:
                pass
        
            try:
                step = int(self.steps[d])
            except:
                step = 0
        
            try:
                data = client.get_date(d)
                meals = data.meals
                notes = data.notes
                totals = data.totals
                carbs = totals['carbohydrates']
                cals = totals['calories']
            except:
                pass
            rec = Record(d,
                         weight=weight,
                         steps=step,
                         carbs=carbs,
                         cals=cals,
                         meals=meals)
            rec.grok_notes(notes)
            self.records.append(rec)
            d += oneday
    
    def dump(self, dest):
        if dest.endswith('.xls') or dest.endswith('.xlsx'):
            self.dump_to_xls(dest)
        elif dest == '-':
            self.dump_to_stream(sys.stdout)
        else:
            with open(dest,'w') as out:
                self.dump_to_stream(out)
                
    def dump_to_stream(self, out):
        print("DATE\tWEIGHT\tSTEPS\tCARBS\tCALS", file=out)        
        for rec in self.records:
            print("{date}\t{weight}\t{steps}\t{carbs}\t{cals}".
                  format(date=rec.date,
                         weight=rec.weight,
                         steps=rec.steps,
                         carbs=rec.carbs,
                         cals=rec.cals),
                  file=out)
            
    def dump_to_xls(self, dest):
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity"
        ws.cell(row=1,column=1, value='DATE')
        ws.cell(row=1,column=2, value='WEIGHT')
        ws.cell(row=1,column=3, value='STEPS')
        ws.cell(row=1,column=4, value='CARBS')
        ws.cell(row=1,column=5, value='CALS')

        ws.cell(row=1,column=6, value='B CARBS')
        ws.cell(row=1,column=7, value='B CALS')
        ws.cell(row=1,column=8, value='L CARBS')
        ws.cell(row=1,column=9, value='L CALS')
        ws.cell(row=1,column=10, value='D CARBS')
        ws.cell(row=1,column=11, value='D CALS')
        ws.cell(row=1,column=12, value='S CARBS')
        ws.cell(row=1,column=13, value='S CALS')

        ws.cell(row=1,column=14, value='NOTES')

        r=2
        for rec in self.records:
            ws.cell(row=r, column=1, value=rec.date)
            ws.cell(row=r, column=2, value=rec.weight)
            ws.cell(row=r, column=3, value=rec.steps)
            ws.cell(row=r, column=4, value=rec.carbs)           
            ws.cell(row=r, column=5, value=rec.cals)
            offset = 6
            inc = 2
            for m in range(0,4): #['breakfast', 'lunch', 'dinner', 'snack']:
                try:
                    meal = rec.meals[m].totals
                    mcarb = meal['carbohydrates']
                    mcals = meal['calories']
                    ws.cell(row=r, column=offset, value=mcarb)
                    ws.cell(row=r, column=offset+1, value=mcals)
                except:
                    pass
                offset += inc
            
            if rec.notes != '':
                cell = ws.cell(row=r, column=offset, value='{}'.format(len(rec.notes.split('\n'))))
                cell.comment = Comment(rec.notes, 'bta')
            offset +=1
            r += 1
        wb.save(dest)
        
if __name__ == "__main__":
    parser = ArgumentParser(description='Better than the alternative.')
    parser.add_argument('username', nargs=1,
                        help='MyFitnessPal login name')
    parser.add_argument('--verbose', '-v', action='count',
                        help='be verbose')
    parser.add_argument('--days', '-d', type=int, default=7,
                        help='Number of days to report on')
    parser.add_argument('--output', '-o', default='-',
                        help='output destination: "-" means printed, something else will write to a file')
    args = parser.parse_args()
    bta = BTA(args.username[0])
    bta.verbose = args.verbose
    nd = args.days
    d2 = date.today() - oneday
    d1 = d2 - nd * oneday
    bta.get_data(d1,d2)
    bta.dump(args.output)
    
    
